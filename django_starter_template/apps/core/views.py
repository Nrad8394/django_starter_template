"""
Core views for the application
"""

from django.http import JsonResponse, HttpResponse
from django.middleware.csrf import get_token
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets, filters
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import (
    extend_schema,
    OpenApiParameter,
    OpenApiResponse,
    OpenApiTypes,
)
import csv
import io
from django.db import transaction
from django.db.models.fields import NOT_PROVIDED
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework import generics, permissions, filters
from django.contrib.admin.models import LogEntry
from .serializers import LogEntrySerializer, SystemSettingsSerializer, FileUploadSerializer
from celery.result import AsyncResult
from drf_spectacular.utils import (
    extend_schema,
    OpenApiParameter,
    OpenApiResponse,
    OpenApiTypes,
)
import django_filters
from .filters import LogEntryFilter
from .schema import common_responses, pagination_parameters
from django.db.models import F, Value, Func
import logging
from datetime import date, timedelta, datetime
from django.utils import timezone


class BaseModelViewSet(viewsets.ModelViewSet):
    """
    Base ModelViewSet with enhanced features including bulk operations, import/export, and statistics.

    This base viewset provides:
    - Standard CRUD operations from ModelViewSet
    - Search and ordering filters
    - Bulk creation endpoint
    - Bulk import from CSV
    - Bulk export to CSV
    - Statistics endpoint

    Subclasses should define:
    - queryset
    - serializer_class (or get_serializer_class)
    - search_fields (optional)
    - ordering_fields (optional)
    - ordering (optional)
    - permission_classes
    """

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    USER_EMAIL_ANNOTATION = staticmethod(
        lambda field_name: Func(
            F(f"{field_name}__email"),
            function="COALESCE",
            template="%(function)s(%(expressions)s, '')",
        )
    )

    def get_queryset(self):
        """
        Automatically annotates CreatedByName and UpdatedByName
        for any queryset where the model has CreatedBy / UpdatedBy fields.
        """
        queryset = super().get_queryset()
        # exclude deleted items if SoftDeleteMixin is used
        # if hasattr(self.queryset.model, "is_deleted"):
        #     queryset = queryset.filter(is_deleted=False)

        model_fields = {field.name for field in self.queryset.model._meta.get_fields()}
        annotations = {}
        select_related_fields = []

        # Handle CreatedBy annotation
        if "created_by" in model_fields:
            annotations["created_by_email"] = self.USER_EMAIL_ANNOTATION("created_by")
            select_related_fields.append("created_by")

        # Handle UpdatedBy annotation
        if "updated_by" in model_fields:
            annotations["updated_by_email"] = self.USER_EMAIL_ANNOTATION("updated_by")
            select_related_fields.append("updated_by")

        # Handle deleted_by annotation
        if "deleted_by" in model_fields:
            annotations["deleted_by_email"] = self.USER_EMAIL_ANNOTATION("deleted_by")
            select_related_fields.append("deleted_by")

        # Only annotate/select if available
        if annotations:
            queryset = queryset.select_related(*select_related_fields).annotate(
                **annotations
            )

        return queryset

    @classmethod
    def get_extra_actions(cls):
        """
        Return a list of extra actions, including inherited ones from base classes.
        """
        actions = super().get_extra_actions()
        # Include actions from base classes
        for base_cls in cls.__mro__[1:]:  # Skip self
            for method_name in dir(base_cls):
                method = getattr(base_cls, method_name, None)
                if method and hasattr(method, "url_path"):
                    if method not in actions:
                        actions.append(method)
        return actions

    @extend_schema(
        summary="Bulk create objects",
        description="Create multiple objects in a single request",
        request={"type": "array", "items": {"type": "object"}},
        responses={
            201: OpenApiResponse(
                description="Objects created successfully",
                response={
                    "type": "object",
                    "properties": {
                        "count": {
                            "type": "integer",
                            "description": "Number of objects created",
                        },
                        "objects": {"type": "array", "items": {"type": "object"}},
                    },
                },
            ),
            **common_responses,
        },
    )
    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        """Create multiple objects in bulk"""
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            instances = serializer.save()

        response_serializer = self.get_serializer(instances, many=True)
        return Response(
            {"count": len(instances), "objects": response_serializer.data},
            status=status.HTTP_201_CREATED,
        )

    @extend_schema(
        summary="Bulk import from CSV or Excel",
        description="Import objects from a CSV or Excel (.xlsx) file",
        request=FileUploadSerializer,
        responses={
            200: OpenApiResponse(
                description="Import completed",
                response={
                    "type": "object",
                    "properties": {
                        "imported": {
                            "type": "integer",
                            "description": "Number of objects imported",
                        },
                        "errors": {"type": "array", "items": {"type": "string"}},
                    },
                },
            ),
            **common_responses,
        },
    )
    @action(detail=False, methods=["post"])
    def bulk_import(self, request):
        """Import objects from CSV or Excel file"""
        if "file" not in request.FILES:
            return Response(
                {"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES["file"]
        file_name = file.name.lower()

        def normalize_header(header):
            """Normalize header to snake_case"""
            if header:
                return header.lower().replace(" ", "_")
            return header

        if file_name.endswith(".csv"):
            # Handle CSV
            file_data = file.read().decode("utf-8")
            reader = csv.reader(io.StringIO(file_data))
            try:
                headers = next(reader)
            except StopIteration:
                return Response({"error": "Empty file"}, status=status.HTTP_400_BAD_REQUEST)
            
            normalized_headers = [normalize_header(h) for h in headers]
            rows = []
            for row in reader:
                if not row: continue
                row_dict = dict(zip(normalized_headers, row))
                rows.append(row_dict)
        elif file_name.endswith(".xlsx"):
            # Handle Excel
            try:
                workbook = load_workbook(file, read_only=True)
                sheet = workbook.active
                rows_iter = sheet.iter_rows(values_only=True)
                headers = next(rows_iter, None)
                if not headers:
                     return Response({"error": "Empty file"}, status=status.HTTP_400_BAD_REQUEST)
                normalized_headers = [normalize_header(h) for h in headers]
                rows = []
                for row in rows_iter:
                    if not any(row): continue
                    row_dict = dict(zip(normalized_headers, row))
                    rows.append(row_dict)
            except Exception as e:
                return Response({"error": f"Invalid Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(
                {"error": "File must be CSV or Excel (.xlsx)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        imported = 0
        errors = []
        results = []

        with transaction.atomic():
            for row_num, row in enumerate(rows, start=2):
                result = {
                    "row": row_num,
                    "data": row,
                    "status": "pending",
                    "errors": None
                }
                try:
                    # Find existing object by unique identifier, e.g., 'id'
                    instance = None
                    if "id" in row and row["id"]:
                        try:
                            instance = self.get_queryset().get(id=row["id"])
                        except self.get_queryset().model.DoesNotExist:
                            instance = None
                    
                    # For bulk import, use the create serializer if it's a new object
                    # This ensures proper handling of related fields like role_name -> role
                    if instance is None:
                        # Temporarily set action to 'create' to get the create serializer
                        old_action = self.action
                        self.action = 'create'
                        serializer = self.get_serializer(data=row)
                        self.action = old_action
                    else:
                        # For updates, use update serializer
                        old_action = self.action
                        self.action = 'update'
                        serializer = self.get_serializer(instance, data=row)
                        self.action = old_action
                    
                    if serializer.is_valid():
                        # Save object
                        saved_instance = serializer.save()
                        
                        # Validate that required fields are actually set
                        # This catches cases where null=False fields might not be validated properly
                        validation_errors = {}
                        for field in self.get_queryset().model._meta.get_fields():
                            if hasattr(field, 'null') and not field.null and not field.primary_key:
                                # Field doesn't allow NULL
                                if getattr(saved_instance, field.name, None) is None:
                                    validation_errors[field.name] = f"This field cannot be empty. {field.help_text}"
                        
                        if validation_errors:
                            # Rollback and report error
                            result["status"] = "error"
                            result["errors"] = validation_errors
                            errors.append(f"Row {row_num}: Missing required field(s) - {validation_errors}")
                        else:
                            imported += 1
                            result["status"] = "success"
                    else:
                        result["status"] = "error"
                        result["errors"] = serializer.errors
                        # Format errors more clearly for the user
                        formatted_errors = self._format_serializer_errors(serializer.errors, row_num)
                        errors.extend(formatted_errors)
                except Exception as e:
                    result["status"] = "error"
                    result["errors"] = str(e)
                    errors.append(f"Row {row_num}: {str(e)}")
                
                results.append(result)

        return Response({
            "imported": imported, 
            "errors": errors,
            "total_rows": len(rows),
            "results": results
        })

    def _format_serializer_errors(self, errors, row_num):
        """Format serializer errors into user-friendly messages"""
        formatted = []
        for field, messages in errors.items():
            if isinstance(messages, list):
                for msg in messages:
                    formatted.append(f"Row {row_num}, Column '{field}': {msg}")
            elif isinstance(messages, dict):
                # Nested errors
                for subfield, submsg in messages.items():
                    formatted.append(f"Row {row_num}, Column '{field}.{subfield}': {submsg}")
            else:
                formatted.append(f"Row {row_num}, Column '{field}': {messages}")
        return formatted

    # bulk import template action to return a sample CSV/Excel file
    @extend_schema(
        summary="Get bulk import template",
        description="Download a sample CSV or Excel template for bulk import",
        parameters=[
            OpenApiParameter(
                name="format",
                type=OpenApiTypes.STR,
                description="Template format: 'csv' or 'xlsx' (default: 'xlsx')",
                default="xlsx",
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="CSV or Excel template file",
                response={"type": "string", "format": "binary"},
            ),
            **common_responses,
        },
    )
    @extend_schema(
        summary="Bulk delete objects",
        description="Delete multiple objects by ID in a single request",
        request={"type": "object", "properties": {"ids": {"type": "array", "items": {"type": "integer"}, "description": "List of object IDs to delete"}}},
        responses={
            200: OpenApiResponse(
                description="Deletion successful",
                response={"type": "object", "properties": {"count": {"type": "integer", "description": "Number of objects deleted"}}},
            ),
            **common_responses,
        },
    )
    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """Delete multiple objects in bulk"""
        ids = request.data.get("ids", [])
        
        if not ids:
            return Response(
                {"error": "No IDs provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        if not isinstance(ids, list):
            return Response(
                {"error": "IDs must be a list"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Get the queryset to delete (respecting permissions)
        queryset = self.filter_queryset(self.get_queryset()).filter(id__in=ids)
        count, _ = queryset.delete()
        
        return Response(
            {"count": count, "message": f"Successfully deleted {count} object(s)"},
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"])
    def bulk_import_template(self, request):
        """Download a sample CSV or Excel template for bulk import"""
        export_format = request.query_params.get("export_format", "xlsx").lower()

        # Get model fields for the template
        model = self.queryset.model

        # Determine required fields and optional recommended fields
        required_fields = []
        optional_fields = []
        field_help_text = {}  # Store help text for each field
        
        old_action = getattr(self, "action", None)
        try:
            # Ask for the create serializer so template matches API validation
            self.action = "create"
            serializer_class = self.get_serializer_class()
            try:
                serializer = serializer_class()
                
                # Separate required vs optional fields
                for name, fld in serializer.fields.items():
                    if getattr(fld, "read_only", False):
                        continue  # Skip read-only fields
                    
                    # Get help text from serializer field first, fall back to model
                    help_text = getattr(fld, "help_text", None) or ""
                    if not help_text:
                        try:
                            model_field = model._meta.get_field(name)
                            help_text = getattr(model_field, "help_text", "") or ""
                        except Exception:
                            help_text = ""
                    
                    field_help_text[name] = help_text
                
            except Exception:
                # If serializer instantiation fails, fall back to model-based approach
                required_fields = []
                optional_fields = []
        finally:
            self.action = old_action
        
        # Combine: required fields first, then optional recommended fields
        fields = required_fields + optional_fields
        
        # Normalize headers to title case
        headers = [field.replace("_", " ").title() for field in fields]

        if export_format == "csv":
            # CSV template
            output = io.StringIO()
            writer = csv.writer(output)

            # Write headers on row 1
            writer.writerow(headers)

            # Blank row for data to start (row 2)
            writer.writerow([])

            # Leave a small spacer, then append instructions and numbered helptext at the very end
            writer.writerow([])
            writer.writerow(["INSTRUCTIONS:"])
            writer.writerow(["REQUIRED fields must be filled in. OPTIONAL fields are recommended for better data organization."])
            writer.writerow([])
            
            # Numbered helptext lines
            for idx, f in enumerate(fields, start=1):
                is_required = f in required_fields
                req_marker = "[REQUIRED]" if is_required else "[OPTIONAL]"
                ht = field_help_text.get(f, "")
                writer.writerow([f"{idx}. {f.replace('_', ' ').title()} {req_marker}: {str(ht)}"])

            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type="text/csv")
            filename = f"{self.queryset.model._meta.verbose_name.title().replace(' ', '_')}_Import_Template.csv"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
        elif export_format == "xlsx":
            # Excel template
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = self.queryset.model._meta.model_name

            # Write headers on row 1
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # Leave row 2 for data to start
            # Place instructions and numbered helptext near the end of the sheet (after a small spacer)
            instr_row = 10
            sheet.cell(row=instr_row, column=1, value="INSTRUCTIONS:")
            sheet.cell(row=instr_row + 1, column=1, value="REQUIRED fields must be filled in. OPTIONAL fields are recommended for better data organization.")
            sheet.cell(row=instr_row + 2, column=1, value="")

            # Write numbered helptext starting on the row after instructions
            for idx, field_name in enumerate(fields, start=1):
                is_required = field_name in required_fields
                req_marker = "[REQUIRED]" if is_required else "[OPTIONAL]"
                help_text = field_help_text.get(field_name, "")
                sheet.cell(row=instr_row + 2 + idx, column=1, value=str(f"{idx}. {field_name.replace('_', ' ').title()} {req_marker}: {help_text}"))
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            filename = f"{self.queryset.model._meta.verbose_name.title().replace(' ', '_')}_Import_Template.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
        else:
            return Response(
                {"error": 'Invalid format. Use "csv" or "xlsx"'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return response

    @extend_schema(
        summary="Bulk export to CSV or Excel",
        description="Export all objects to CSV or Excel format",
        parameters=[
            OpenApiParameter(
                name="fields",
                type=OpenApiTypes.STR,
                description="Comma-separated list of fields to export",
            ),
            OpenApiParameter(
                name="format",
                type=OpenApiTypes.STR,
                description="Export format: 'csv' or 'xlsx' (default: 'xlsx')",
                default="xlsx",
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="CSV or Excel file",
                response={"type": "string", "format": "binary"},
            ),
            **common_responses,
        },
    )
    @action(detail=False, methods=["get"])
    def bulk_export(self, request):
        """Export objects to CSV or Excel"""
        queryset = self.filter_queryset(self.get_queryset())
        model = queryset.model
        fields = request.query_params.get("fields", None)
        export_format = request.query_params.get("file_format", "xlsx").lower()

        if fields:
            fields = [f.strip() for f in fields.split(",")]
        else:
            # Default to model fields
            fields = [f.name for f in model._meta.fields]

        # Normalize headers to title case
        headers = [field.replace("_", " ").title() for field in fields]

        if export_format == "csv":
            # CSV export — headers on row 1, data starting row 2, instructions and numbered helptext appended at the end
            output = io.StringIO()
            writer = csv.writer(output)

            # Write headers on row 1
            writer.writerow(headers)

            # Write data rows
            for obj in queryset:
                row = []
                for field in fields:
                    value = getattr(obj, field, "")
                    # Handle callable values (methods, properties, related managers)
                    if callable(value):
                        try:
                            value = value()
                        except TypeError as e:
                            # Handle many-to-many managers or other callables that need special handling
                            if "manager" in str(e):
                                # For many-to-many relationships, get the related objects
                                if hasattr(value, 'all'):
                                    value = ", ".join([str(item) for item in value.all()])
                                else:
                                    value = ""
                            else:
                                value = ""
                    row.append(str(value))
                writer.writerow(row)



            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = (
                f'attachment; filename="{self.queryset.model._meta.model_name}.csv"'
            )
        elif export_format == "xlsx":
            # Excel export
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = self.queryset.model._meta.model_name


            # Headers on row 1, data starting from row 2. Append instructions and numbered helptext after data.
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # Write data starting from row 2
            for row_num, obj in enumerate(queryset, 2):
                for col_num, field in enumerate(fields, 1):
                    value = getattr(obj, field, "")
                    # Handle callable values (methods, properties, related managers)
                    if callable(value):
                        try:
                            value = value()
                        except TypeError as e:
                            # Handle many-to-many managers or other callables that need special handling
                            if "manager" in str(e):
                                # For many-to-many relationships, get the related objects
                                if hasattr(value, 'all'):
                                    value = ", ".join([str(item) for item in value.all()])
                                else:
                                    value = ""
                            else:
                                value = ""
                    sheet.cell(row=row_num, column=col_num, value=str(value))



            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                f'attachment; filename="{self.queryset.model._meta.model_name}.xlsx"'
            )
        else:
            return Response(
                {"error": 'Invalid format. Use "csv" or "xlsx"'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return response

    @extend_schema(
        summary="Get statistics",
        description="Get basic statistics for the model",
        responses={
            200: OpenApiResponse(
                description="Model statistics",
                response={
                    "type": "object",
                    "properties": {
                        "total_count": {
                            "type": "integer",
                            "description": "Total number of objects",
                        },
                        "filtered_count": {
                            "type": "integer",
                            "description": "Number of objects after filtering",
                        },
                    },
                },
            ),
            **common_responses,
        },
    )
    @action(detail=False, methods=["get"])
    def statistics(self, request):
        """Get basic statistics"""
        queryset = self.get_queryset()
        filtered_queryset = self.filter_queryset(queryset)

        stats = {
            "total_count": queryset.count(),
            "filtered_count": filtered_queryset.count(),
        }

        # Allow subclasses to add more stats
        extra_stats = self.get_extra_statistics(request, queryset, filtered_queryset)
        stats.update(extra_stats)

        return Response(stats)

    def get_extra_statistics(self, request, queryset, filtered_queryset):
        """
        Override this method in subclasses to add custom statistics.
        Should return a dict of additional statistics.
        """
        return {}


@extend_schema(
    summary="Health check endpoint",
    description="Simple health check to verify the API is running",
    responses={
        200: OpenApiResponse(
            description="API is healthy",
            response={
                "type": "object",
                "properties": {
                    "status": {"type": "string", "example": "healthy"},
                    "message": {"type": "string", "example": "API is running"},
                },
            },
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET"])
@permission_classes([AllowAny])
def health_check(request):
    """
    Health check endpoint for monitoring and load balancers.
    """
    return Response(
        {"status": "healthy", "message": "API is running"}, status=status.HTTP_200_OK
    )


@extend_schema(
    summary="Get CSRF token",
    description="Get CSRF token for frontend applications that need session-based authentication",
    responses={
        200: OpenApiResponse(
            description="CSRF token",
            response={
                "type": "object",
                "properties": {
                    "csrfToken": {
                        "type": "string",
                        "description": "CSRF token for session authentication",
                    },
                },
            },
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
@ensure_csrf_cookie
def get_csrf_token(request):
    """
    Get CSRF token for frontend applications.

    This endpoint ensures a CSRF cookie is set and returns the token
    for frontend applications that need to make session-authenticated requests.
    """
    csrf_token = get_token(request)
    return Response({"csrfToken": csrf_token}, status=status.HTTP_200_OK)


@require_http_methods(["GET"])
@ensure_csrf_cookie
def csrf_token_view(request):
    """
    Alternative CSRF token endpoint for non-DRF clients.
    Returns CSRF token as JSON.
    """
    csrf_token = get_token(request)
    return JsonResponse({"csrfToken": csrf_token})


@extend_schema(
    summary="List history items (audit log)",
    description="Returns a paginated list of history items (audit log entries) for all models in the system. Only staff/admins can view.",
    parameters=[
        OpenApiParameter(
            name="user",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.STR,
            description="Filter by username (icontains)",
        ),
        OpenApiParameter(
            name="user_id",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.INT,
            description="Filter by user ID",
        ),
        OpenApiParameter(
            name="action_time_after",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.DATETIME,
            description="Filter by action time after",
        ),
        OpenApiParameter(
            name="action_time_before",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.DATETIME,
            description="Filter by action time before",
        ),
        OpenApiParameter(
            name="action_flag",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.INT,
            description="Filter by action flag (1=Addition, 2=Change, 3=Deletion)",
        ),
        OpenApiParameter(
            name="content_type",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.STR,
            description="Filter by model name (icontains)",
        ),
        OpenApiParameter(
            name="object_id",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.STR,
            description="Filter by object ID (icontains)",
        ),
        OpenApiParameter(
            name="object_repr",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.STR,
            description="Filter by object representation (icontains)",
        ),
        OpenApiParameter(
            name="change_message",
            location=OpenApiParameter.QUERY,
            type=OpenApiTypes.STR,
            description="Filter by change message (icontains)",
        ),
    ]
    + pagination_parameters,
    responses={200: LogEntrySerializer(many=True), **common_responses},
    tags=["Core"],
)
class HistoryListView(generics.ListAPIView):
    queryset = LogEntry.objects.all().order_by("-action_time")
    serializer_class = LogEntrySerializer
    filter_backends = [
        django_filters.rest_framework.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = LogEntryFilter
    search_fields = [
        "user__username",
        "object_repr",
        "change_message",
        "content_type__model",
    ]
    ordering_fields = [
        "action_time",
        "user__username",
        "content_type__model",
        "action_flag",
    ]
    ordering = ["-action_time"]
    permission_classes = [permissions.IsAdminUser]


@extend_schema(
    summary="Get task status",
    description="Check the status of an asynchronous Celery task by its task ID",
    parameters=[
        OpenApiParameter(
            name="task_id",
            location=OpenApiParameter.PATH,
            type=OpenApiTypes.UUID,
            description="The UUID of the Celery task to check",
            required=True,
        )
    ],
    responses={
        200: OpenApiResponse(
            description="Task status information",
            response={
                "type": "object",
                "properties": {
                    "task_id": {
                        "type": "string",
                        "format": "uuid",
                        "description": "The task ID",
                    },
                    "status": {
                        "type": "string",
                        "enum": [
                            "PENDING",
                            "PROGRESS",
                            "SUCCESS",
                            "FAILURE",
                            "RETRY",
                            "REVOKED",
                        ],
                        "description": "Current task status",
                    },
                    "result": {
                        "type": "object",
                        "description": "Task result (only present when status is SUCCESS)",
                    },
                    "error": {
                        "type": "string",
                        "description": "Error message (only present when status is FAILURE)",
                    },
                    "traceback": {
                        "type": "string",
                        "description": "Error traceback (only present when status is FAILURE)",
                    },
                    "date_done": {
                        "type": "string",
                        "format": "date-time",
                        "description": "When the task completed (only present when task is done)",
                    },
                },
            },
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def task_status(request, task_id):
    """
    Get the status of a Celery task by its ID.

    This endpoint allows clients to check the progress and result of
    asynchronous tasks like AI chat responses, document generation, etc.
    """
    # task_id is already validated as UUID by Django's URL converter
    # No need to validate again since <uuid:task_id> ensures it's a valid UUID

    # Get task result from Celery
    task_result = AsyncResult(str(task_id))

    if task_result.state == "PENDING":
        # Task is waiting to be processed
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
        }
    elif task_result.state == "PROGRESS":
        # Task is in progress (if using progress tracking)
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "current": getattr(task_result.info, "current", None),
            "total": getattr(task_result.info, "total", None),
            "message": getattr(task_result.info, "message", ""),
        }
    elif task_result.state == "SUCCESS":
        # Task completed successfully
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "result": task_result.result,
            "date_done": (
                task_result.date_done.isoformat() if task_result.date_done else None
            ),
        }
    elif task_result.state == "FAILURE":
        # Task failed
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "error": str(task_result.info),
            "traceback": getattr(task_result.info, "traceback", None),
            "date_done": (
                task_result.date_done.isoformat() if task_result.date_done else None
            ),
        }
    else:
        # Other states (RETRY, REVOKED, etc.)
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "info": str(task_result.info) if task_result.info else None,
            "date_done": (
                task_result.date_done.isoformat() if task_result.date_done else None
            ),
        }

    return Response(response_data, status=status.HTTP_200_OK)

@extend_schema(
    summary="Pending Approvals",
    description="""
    Get items awaiting approval for the current user.
    Scoped to user's role and permissions.
    """,
    responses={
        200: OpenApiResponse(
            description="Pending approvals",
            response={"type": "array", "items": {"type": "object"}},
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def pending_approvals(request):
    """
    Get pending approvals for the current user.
    This is a placeholder - implement based on your specific approval workflow.
    """
    user = request.user
    approvals = []

    # Example: Check for users awaiting approval (admin/registrar)
    if user.is_superuser or (
        hasattr(user, "role") and user.role and user.role.name in ["admin", "registrar"]
    ):
        try:
            from apps.accounts.models import User

            pending_users = User.objects.filter(
                is_approved=False, is_active=True
            ).order_by("-created_at")[:10]

            for u in pending_users:
                approvals.append(
                    {
                        "id": str(u.id),
                        "type": "user_approval",
                        "title": f"Approve user: {u.get_full_name()}",
                        "description": f'{u.email} - {u.role.display_name if u.role else "No role"}',
                        "created_at": (
                            u.created_at.isoformat()
                            if hasattr(u, "created_at")
                            else None
                        ),
                    }
                )
        except Exception as e:
            print(f"Error fetching pending approvals: {e}")

    # Add more approval types based on your workflow
    # - Enrollment approvals
    # - Leave requests
    # - Resource booking approvals
    # etc.

    return Response(approvals)


@extend_schema(
    summary="User Notifications",
    description="Get notifications for the current user",
    parameters=[
        OpenApiParameter(
            name="unread_only",
            type=OpenApiTypes.BOOL,
            description="Filter for unread notifications only",
        ),
    ],
    responses={
        200: OpenApiResponse(
            description="User notifications",
            response={"type": "array", "items": {"type": "object"}},
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_notifications(request):
    """
    Get notifications for the current user.
    This is a placeholder - implement based on your notifications system.
    """
    user = request.user
    unread_only = request.query_params.get("unread_only", "false").lower() == "true"
    notifications = []

    # Check if notifications app is available
    try:
        from apps.notifications.models import Notification

        queryset = Notification.objects.filter(recipient=user)

        if unread_only:
            queryset = queryset.filter(is_read=False)

        queryset = queryset.order_by("-created_at")[:50]

        for notif in queryset:
            notifications.append(
                {
                    "id": str(notif.id),
                    "title": notif.title if hasattr(notif, "title") else "Notification",
                    "message": (
                        notif.message if hasattr(notif, "message") else str(notif)
                    ),
                    "type": (
                        notif.notification_type
                        if hasattr(notif, "notification_type")
                        else "info"
                    ),
                    "is_read": notif.is_read if hasattr(notif, "is_read") else False,
                    "created_at": (
                        notif.created_at.isoformat()
                        if hasattr(notif, "created_at")
                        else None
                    ),
                }
            )
    except ImportError:
        # Notifications app not available, return sample data
        pass

    return Response(notifications)


@extend_schema(
    summary="Mark Notification as Read",
    description="Mark a notification as read",
    responses={
        200: OpenApiResponse(description="Notification marked as read"),
        404: OpenApiResponse(description="Notification not found"),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mark_notification_read(request, notification_id):
    """
    Mark a specific notification as read.
    """
    user = request.user

    try:
        from apps.notifications.models import Notification

        notification = Notification.objects.get(id=notification_id, recipient=user)
        notification.is_read = True
        notification.save()

        return Response({"status": "success", "message": "Notification marked as read"})
    except ImportError:
        return Response(
            {"error": "Notifications system not available"},
            status=status.HTTP_501_NOT_IMPLEMENTED,
        )
    except Notification.DoesNotExist:
        return Response(
            {"error": "Notification not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    summary="Get System Settings",
    description="Get current system-wide configuration settings. Admin only.",
    responses={
        200: OpenApiResponse(
            description="System settings", response={"type": "object"}
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_system_settings(request):
    """
    Get current system settings.
    Only administrators can view settings.
    """
    from apps.core.models import SystemSettings
    from apps.core.serializers import SystemSettingsSerializer

    user = request.user

    # Check if user is admin
    if not (
        user.is_superuser
        or (hasattr(user, "role") and user.role and user.role.name == "admin")
    ):
        return Response(
            {"error": "Only administrators can view system settings"},
            status=status.HTTP_403_FORBIDDEN,
        )

    # Load settings (singleton)
    settings_obj = SystemSettings.load()
    serializer = SystemSettingsSerializer(settings_obj)

    return Response(serializer.data)


@extend_schema(
    summary="Update System Settings",
    description="Update system-wide configuration settings. Admin only.",
    request=SystemSettingsSerializer,
    responses={
        200: OpenApiResponse(
            description="Settings updated successfully", response={"type": "object"}
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def update_system_settings(request):
    """
    Update system settings.
    Only administrators can modify settings.
    """
    from apps.core.models import SystemSettings
    from apps.core.serializers import SystemSettingsSerializer

    user = request.user

    # Check if user is admin
    if not (
        user.is_superuser
        or (hasattr(user, "role") and user.role and user.role.name == "admin")
    ):
        return Response(
            {"error": "Only administrators can modify system settings"},
            status=status.HTTP_403_FORBIDDEN,
        )

    # Load settings (singleton)
    settings_obj = SystemSettings.load()

    # Validate input
    serializer = SystemSettingsSerializer(data=request.data, partial=True)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Update fields
    for field, value in serializer.validated_data.items():
        if field not in ["updated_at", "updated_by_email"]:  # Skip read-only fields
            setattr(settings_obj, field, value)

    # Set updated_by
    settings_obj.updated_by = user
    settings_obj.save()

    # Return updated settings
    response_serializer = SystemSettingsSerializer(settings_obj)
    return Response(response_serializer.data)


@extend_schema(
    summary="System Settings (Combined GET/PATCH)",
    description="Get or update system settings. Admin only.",
    request=SystemSettingsSerializer,
    responses={
        200: OpenApiResponse(
            description="System settings", response={"type": "object"}
        ),
        **common_responses,
    },
    tags=["Core"],
)
@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def system_settings(request):
    """
    Combined endpoint for system settings.
    GET: Retrieve settings
    PATCH: Update settings
    Admin only.
    """
    # The incoming `request` here may be a DRF `Request` (when routed through
    # DRF machinery). The decorated view functions `get_system_settings` and
    # `update_system_settings` expect a raw Django `HttpRequest` (they are
    # wrapped by `@api_view`) — passing a DRF `Request` directly into them
    # causes DRF to attempt to wrap it again and assert. To avoid that, pass
    # the underlying HttpRequest when available.
    base_request = getattr(request, "_request", request)

    if request.method == "GET":
        return get_system_settings(base_request)
    elif request.method == "PATCH":
        return update_system_settings(base_request)

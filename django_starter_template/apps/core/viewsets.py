"""
``BaseModelViewSet`` — a ModelViewSet with the endpoints every admin UI needs.

On top of DRF's CRUD it adds bulk create / update / delete / restore, CSV and
XLSX import and export, soft-delete-aware listing, audit-field population, and
a statistics hook. A typical resource becomes::

    class ProductViewSet(BaseModelViewSet):
        queryset = Product.objects.all()
        serializer_class = ProductSerializer
        search_fields = ["name", "sku"]
        ordering_fields = ["name", "created_at"]
        export_fields = ["sku", "name", "price", "created_at"]

…and the frontend's generic table component gets working bulk actions, an
export button, and a trash view with no per-resource backend code.

Read before adopting
--------------------
Three deliberate departures from the implementation this was distilled from,
each of which was a live problem:

1. **Bulk update goes through the serializer.** The original had a fast path
   that took ``{"ids": [...], "data": {...}}`` and called
   ``queryset.update(**data)``. ``QuerySet.update()`` writes columns straight
   to SQL: no serializer validation, no ``clean()``, no signals, no
   ``updated_at``, no ``updated_by``, and no restriction on which fields the
   caller may name. A client could set any column on any row it could list,
   including ``is_deleted``, a price, or a foreign key to another tenant's
   record. It is gone. ``bulk_update`` accepts only the per-object form and
   validates each one.

2. **Export requires an explicit field list.** The original defaulted to every
   concrete field on the model, so adding a column to a model silently added
   it to a user-downloadable spreadsheet. Set ``export_fields`` to opt in.

3. **Export is capped.** ``bulk_export`` streamed an unbounded queryset into
   an in-memory workbook; on a large table that is an out-of-memory kill.
   ``max_export_rows`` bounds it and the response says so when it truncates.

Everything else — the soft-delete query modes, the audit annotations, the
import template generator — is carried over largely intact because it earned
its place.
"""

from __future__ import annotations

import csv
import io
import logging

from django.core.exceptions import ImproperlyConfigured
from django.db import transaction
from django.db.models import F, Value
from django.db.models.deletion import ProtectedError, RestrictedError
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, OpenApiResponse, extend_schema
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response

from .bulk_serializers import (
    BulkIdListSerializer,
    BulkUpdateSerializer,
    FileUploadSerializer,
)

logger = logging.getLogger(__name__)

TRUTHY = {"1", "true", "yes", "y", "on"}
FALSY = {"0", "false", "no", "n", "off"}

#: openpyxl is only needed for the XLSX paths. Import it lazily so a project
#: that only ever exports CSV does not have to carry the dependency.
try:  # pragma: no cover - import guard
    from openpyxl import Workbook, load_workbook

    XLSX_AVAILABLE = True
except ImportError:  # pragma: no cover - import guard
    Workbook = load_workbook = None
    XLSX_AVAILABLE = False


class SoftDeleteQueryMixin:
    """
    Lets a client ask for active rows, deleted rows, or both.

    ``?is_deleted=false``  (default) active only
    ``?is_deleted=true``             deleted only — the "Trash" view
    ``?show_deleted=true``           both

    Only ``is_deleted`` and ``show_deleted`` are supported. The source
    implementation also accepted ``ondelete`` as an undocumented third alias
    for the same thing; three spellings of one parameter is three things to
    keep working and three ways for a client to be subtly wrong.
    """

    def _parse_bool(self, value):
        if value is None:
            return None
        normalized = str(value).strip().lower()
        if normalized in TRUTHY:
            return True
        if normalized in FALSY:
            return False
        return None

    def get_soft_delete_mode(self):
        """Return ``True`` (deleted only), ``False`` (active only), or ``"all"``."""
        override = getattr(self, "_soft_delete_mode_override", None)
        if override is not None:
            return override

        params = getattr(self.request, "query_params", {}) or {}

        if "is_deleted" in params:
            parsed = self._parse_bool(params.get("is_deleted"))
            if parsed is not None:
                return parsed

        if "show_deleted" in params:
            if self._parse_bool(params.get("show_deleted")) is True:
                return "all"
            return False

        return False

    def apply_soft_delete_filter(self, queryset):
        model = queryset.model
        if not hasattr(model, "is_deleted"):
            return queryset

        mode = self.get_soft_delete_mode()

        # `objects` hides deleted rows, so anything but "active only" has to
        # start again from the unfiltered manager.
        if mode in (True, "all") and hasattr(model, "all_objects"):
            queryset = model.all_objects.all()

        if mode is True:
            return queryset.filter(is_deleted=True)
        if mode is False:
            return queryset.filter(is_deleted=False)
        return queryset

    def get_queryset_including_deleted(self):
        """
        The view's own queryset — role scoping and all — plus deleted rows.

        Temporarily flips the mode rather than rebuilding the queryset, so any
        ``get_queryset()`` override a subclass has written (tenant filtering,
        ``select_related``, permission scoping) still applies. Rebuilding from
        ``Model.all_objects`` directly here is the tempting shortcut and it
        drops exactly that scoping — which turns a Trash view into a data
        leak across tenants.
        """
        previous = getattr(self, "_soft_delete_mode_override", None)
        self._soft_delete_mode_override = "all"
        try:
            return self.get_queryset()
        finally:
            self._soft_delete_mode_override = previous

    def get_object_including_deleted(self):
        """``get_object()``, but soft-deleted rows are visible."""
        queryset = self.filter_queryset(self.get_queryset_including_deleted())
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field
        lookup_value = self.kwargs.get(lookup_url_kwarg)

        obj = queryset.get(**{self.lookup_field: lookup_value})
        self.check_object_permissions(self.request, obj)
        return obj


class BaseModelViewSet(SoftDeleteQueryMixin, viewsets.ModelViewSet):
    """
    Configure by setting these on the subclass:

    ``queryset`` / ``serializer_class``
        As for any ModelViewSet.
    ``search_fields`` / ``ordering_fields`` / ``filterset_fields``
        Passed to the filter backends.
    ``export_fields``
        Field names allowed in CSV/XLSX export. Empty disables export.
    ``import_fields``
        Field names accepted by import. Empty means "whatever the create
        serializer accepts".
    ``max_export_rows``
        Row cap for a single export.
    ``bulk_operations_enabled``
        Set ``False`` to remove all bulk endpoints from a resource.
    """

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    export_fields: list[str] = []
    import_fields: list[str] = []
    max_export_rows: int = 10_000
    max_bulk_rows: int = 1_000
    bulk_operations_enabled: bool = True

    # ------------------------------------------------------------------
    # Queryset construction
    # ------------------------------------------------------------------

    def is_schema_request(self) -> bool:
        """True while drf-spectacular is introspecting rather than serving."""
        return bool(getattr(self, "swagger_fake_view", False))

    def get_model_class(self):
        """
        Resolve the model even when ``queryset`` is not a class attribute.

        Views that build their queryset entirely in ``get_queryset()`` are
        common; the bulk endpoints still need the model.
        """
        queryset = getattr(self, "queryset", None)
        if queryset is not None and getattr(queryset, "model", None) is not None:
            return queryset.model

        try:
            model = getattr(self.get_queryset(), "model", None)
            if model is not None:
                return model
        except Exception:  # noqa: BLE001 - best-effort resolution
            logger.debug("get_queryset() failed during model resolution", exc_info=True)

        serializer_class = self.get_serializer_class()
        model = getattr(getattr(serializer_class, "Meta", None), "model", None)
        if model is not None:
            return model

        raise ImproperlyConfigured(
            f"{self.__class__.__name__} cannot resolve its model. Set "
            f"`queryset`, or give the serializer a `Meta.model`."
        )

    def get_queryset(self):
        """
        Apply soft-delete visibility and annotate audit-user emails.

        The annotations mean a list response can show "created by
        alice@example.com" without the serializer declaring a nested user
        object — one JOIN instead of N+1 queries, and no extra payload beyond
        the string. Serializers opt in by declaring
        ``created_by_email = serializers.CharField(read_only=True)``.
        """
        if self.is_schema_request():
            queryset = getattr(self, "queryset", None)
            return queryset.none() if queryset is not None else super().get_queryset()

        queryset = super().get_queryset()
        queryset = self.apply_soft_delete_filter(queryset)

        model_fields = {field.name for field in queryset.model._meta.get_fields()}
        annotations = {}
        related = []

        for field_name in ("created_by", "updated_by", "deleted_by"):
            if field_name in model_fields:
                annotations[f"{field_name}_email"] = Coalesce(
                    F(f"{field_name}__email"), Value("")
                )
                related.append(field_name)

        if annotations:
            queryset = queryset.select_related(*related).annotate(**annotations)

        return queryset

    # ------------------------------------------------------------------
    # Audit fields
    # ------------------------------------------------------------------

    def _current_user(self):
        user = getattr(self.request, "user", None)
        if user is not None and user.is_authenticated:
            return user
        return None

    def perform_create(self, serializer):
        """
        Stamp ``created_by``/``updated_by`` from the request.

        Doing it here rather than in a thread-local "current user" middleware
        is a deliberate choice. Thread-locals make the write path implicit and
        break in the places that matter most — Celery tasks, management
        commands, and async views all have no request, so the field silently
        lands as NULL and nobody notices until an audit. Passing the user
        explicitly keeps the data flow visible and testable.
        """
        model_fields = {f.name for f in serializer.Meta.model._meta.get_fields()}
        extra = {}
        user = self._current_user()
        if user is not None:
            if "created_by" in model_fields:
                extra["created_by"] = user
            if "updated_by" in model_fields:
                extra["updated_by"] = user
        serializer.save(**extra)

    def perform_update(self, serializer):
        model_fields = {f.name for f in serializer.Meta.model._meta.get_fields()}
        extra = {}
        user = self._current_user()
        if user is not None and "updated_by" in model_fields:
            extra["updated_by"] = user
        serializer.save(**extra)

    def perform_destroy(self, instance):
        """Route DELETE through soft delete when the model supports it."""
        if hasattr(instance, "is_deleted"):
            self.validate_soft_delete(instance)
            instance.delete(deleted_by=self._current_user())
            return
        super().perform_destroy(instance)

    # ------------------------------------------------------------------
    # Subclass hooks
    # ------------------------------------------------------------------

    def validate_soft_delete(self, instance):
        """Raise ``ValidationError`` to block a delete on business grounds."""

    def validate_hard_delete(self, instance):
        """Raise ``ValidationError`` to block a permanent delete."""

    def get_extra_statistics(self, request, queryset, filtered_queryset) -> dict:
        """Extra keys for the ``statistics`` endpoint."""
        return {}

    # ------------------------------------------------------------------
    # Guards
    # ------------------------------------------------------------------

    def _require_bulk_enabled(self):
        if not self.bulk_operations_enabled:
            raise PermissionDenied("Bulk operations are disabled for this resource.")

    def _validated_ids(self, request):
        serializer = BulkIdListSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        ids = serializer.validated_data["ids"]
        if len(ids) > self.max_bulk_rows:
            raise ValidationError(
                {"ids": [f"At most {self.max_bulk_rows} ids per request."]}
            )
        return ids

    def _require_soft_delete_support(self, model):
        if not hasattr(model, "is_deleted"):
            raise ValidationError(
                {
                    "non_field_errors": [
                        "This resource does not support soft deletion."
                    ]
                }
            )

    # ------------------------------------------------------------------
    # Bulk create
    # ------------------------------------------------------------------

    def get_serializer(self, *args, **kwargs):
        if getattr(self, "action", None) == "bulk_create":
            kwargs.setdefault("many", True)
        return super().get_serializer(*args, **kwargs)

    @extend_schema(
        summary="Create several objects",
        responses={201: OpenApiResponse(description="Objects created")},
    )
    @action(detail=False, methods=["post"], url_path="bulk_create")
    def bulk_create(self, request):
        self._require_bulk_enabled()

        if isinstance(request.data, list) and len(request.data) > self.max_bulk_rows:
            raise ValidationError(
                f"At most {self.max_bulk_rows} objects per request."
            )

        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            instances = serializer.save(**self._audit_kwargs_for_create())

        return Response(
            {
                "count": len(instances),
                "objects": self.get_serializer(instances, many=True).data,
            },
            status=status.HTTP_201_CREATED,
        )

    def _audit_kwargs_for_create(self):
        model = self.get_model_class()
        model_fields = {f.name for f in model._meta.get_fields()}
        user = self._current_user()
        if user is None:
            return {}
        return {
            field: user
            for field in ("created_by", "updated_by")
            if field in model_fields
        }

    # ------------------------------------------------------------------
    # Bulk update
    # ------------------------------------------------------------------

    @extend_schema(
        summary="Update several objects",
        description=(
            "Body: `{\"updates\": [{\"id\": ..., \"data\": {...}}, ...]}`. "
            "Each entry is validated by the resource serializer, so the same "
            "rules apply as for a single PATCH. All entries succeed or none do."
        ),
        request=BulkUpdateSerializer,
        responses={200: OpenApiResponse(description="Objects updated")},
    )
    @action(detail=False, methods=["post"], url_path="bulk_update")
    def bulk_update(self, request):
        self._require_bulk_enabled()

        payload = BulkUpdateSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        updates = payload.validated_data["updates"]

        if len(updates) > self.max_bulk_rows:
            raise ValidationError(
                {"updates": [f"At most {self.max_bulk_rows} updates per request."]}
            )

        # Resolve every target through the *filtered* queryset, so a caller can
        # only touch rows they can already see. Looking objects up by bare
        # primary key here would let anyone update anything whose id they can
        # guess.
        ids = [entry["id"] for entry in updates]
        visible = self.filter_queryset(self.get_queryset()).filter(pk__in=ids)
        by_id = {str(obj.pk): obj for obj in visible}

        missing = [str(pk) for pk in ids if str(pk) not in by_id]
        if missing:
            raise ValidationError(
                {
                    "updates": [
                        f"No accessible object for id(s): {', '.join(missing)}."
                    ]
                }
            )

        updated = []
        with transaction.atomic():
            for entry in updates:
                instance = by_id[str(entry["id"])]
                serializer = self.get_serializer(
                    instance, data=entry["data"], partial=True
                )
                serializer.is_valid(raise_exception=True)
                self.perform_update(serializer)
                updated.append(serializer.instance)

        return Response(
            {
                "count": len(updated),
                "objects": self.get_serializer(updated, many=True).data,
            }
        )

    # ------------------------------------------------------------------
    # Bulk delete / restore
    # ------------------------------------------------------------------

    @extend_schema(
        summary="Delete several objects",
        request=BulkIdListSerializer,
        responses={200: OpenApiResponse(description="Deletion result")},
    )
    @action(detail=False, methods=["post"], url_path="bulk_delete")
    def bulk_delete(self, request):
        self._require_bulk_enabled()
        ids = self._validated_ids(request)

        model = self.get_model_class()
        queryset = self.filter_queryset(self.get_queryset()).filter(pk__in=ids)

        if not hasattr(model, "is_deleted"):
            try:
                deleted, _ = queryset.delete()
            except (ProtectedError, RestrictedError):
                raise ValidationError(
                    {
                        "non_field_errors": [
                            "One or more records are referenced by other "
                            "records and cannot be deleted."
                        ]
                    }
                )
            return Response({"count": deleted, "blocked": []})

        deleted_by = self._current_user()
        deleted_count = 0
        blocked = []

        # Per-object rather than a single bulk UPDATE, because
        # validate_soft_delete() is a per-object business rule and callers
        # need to know which rows it rejected. The partial-success response
        # below is the reason this cannot be one statement.
        with transaction.atomic():
            for obj in queryset.filter(is_deleted=False):
                try:
                    self.validate_soft_delete(obj)
                    obj.delete(deleted_by=deleted_by)
                    deleted_count += 1
                except (ValidationError, ProtectedError, RestrictedError) as exc:
                    blocked.append({"id": str(obj.pk), "error": str(exc)})

        if blocked and not deleted_count:
            return Response(
                {
                    "count": 0,
                    "blocked": blocked,
                    "message": "Nothing was deleted.",
                },
                status=status.HTTP_409_CONFLICT,
            )

        return Response(
            {
                "count": deleted_count,
                "blocked": blocked,
                "message": f"Deleted {deleted_count} object(s).",
            }
        )

    @extend_schema(
        summary="Restore several soft-deleted objects",
        request=BulkIdListSerializer,
        responses={200: OpenApiResponse(description="Restore result")},
    )
    @action(detail=False, methods=["post"], url_path="bulk_restore")
    def bulk_restore(self, request):
        self._require_bulk_enabled()
        model = self.get_model_class()
        self._require_soft_delete_support(model)

        ids = self._validated_ids(request)

        # Scoped through the view's queryset, not model.all_objects — see the
        # note on get_queryset_including_deleted().
        queryset = self.filter_queryset(self.get_queryset_including_deleted())
        queryset = queryset.filter(pk__in=ids, is_deleted=True)

        restored = 0
        with transaction.atomic():
            for obj in queryset:
                try:
                    self.check_object_permissions(request, obj)
                except PermissionDenied:
                    continue
                obj.restore()
                restored += 1

        return Response(
            {"count": restored, "message": f"Restored {restored} object(s)."}
        )

    @extend_schema(
        summary="Restore a soft-deleted object",
        responses={200: OpenApiResponse(description="Object restored")},
    )
    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        model = self.get_model_class()
        self._require_soft_delete_support(model)

        try:
            instance = self.get_object_including_deleted()
        except model.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if not instance.is_deleted:
            raise ValidationError(
                {"non_field_errors": ["This object is not deleted."]}
            )

        instance.restore()
        return Response(self.get_serializer(instance).data)

    @extend_schema(
        summary="Permanently delete an object",
        description=(
            "Irreversible. The object must already be soft-deleted — this is "
            "a two-step confirmation, not a shortcut past DELETE."
        ),
        responses={204: OpenApiResponse(description="Permanently deleted")},
    )
    @action(detail=True, methods=["post"], url_path="hard_delete")
    def hard_delete(self, request, pk=None):
        model = self.get_model_class()
        self._require_soft_delete_support(model)

        try:
            instance = self.get_object_including_deleted()
        except model.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if not instance.is_deleted:
            raise ValidationError(
                {
                    "non_field_errors": [
                        "Move the object to trash before deleting it "
                        "permanently."
                    ]
                }
            )

        self.validate_hard_delete(instance)
        try:
            instance.hard_delete()
        except (ProtectedError, RestrictedError):
            raise ValidationError(
                {
                    "non_field_errors": [
                        "This record is referenced by other records and "
                        "cannot be permanently deleted."
                    ]
                }
            )

        return Response(status=status.HTTP_204_NO_CONTENT)

    # ------------------------------------------------------------------
    # Import / export
    # ------------------------------------------------------------------

    def get_export_fields(self) -> list[str]:
        """
        Fields a client may export.

        Deliberately empty by default. Defaulting to "every model field" — as
        the implementation this came from did — means every column you add to
        a model later, including internal flags, token hashes and soft-delete
        bookkeeping, silently becomes downloadable. Opt in per resource.
        """
        return list(self.export_fields)

    def _resolve_export_value(self, obj, field_name):
        value = obj
        for part in field_name.split("."):
            value = getattr(value, part, None)
            if value is None:
                return ""

        # Related managers (`obj.tags`) expose .all(); render them as a
        # comma-separated list rather than a repr.
        if hasattr(value, "all"):
            return ", ".join(str(item) for item in value.all())

        if callable(value):
            try:
                value = value()
            except TypeError:
                return ""

        return "" if value is None else str(value)

    @extend_schema(
        summary="Export to CSV or XLSX",
        parameters=[
            OpenApiParameter(
                name="fields",
                type=OpenApiTypes.STR,
                description="Comma-separated subset of the allowed export fields.",
            ),
            OpenApiParameter(
                name="file_format",
                type=OpenApiTypes.STR,
                description="'csv' or 'xlsx' (default 'csv').",
            ),
        ],
        responses={200: OpenApiResponse(description="File download")},
    )
    @action(detail=False, methods=["get"], url_path="bulk_export")
    def bulk_export(self, request):
        allowed = self.get_export_fields()
        if not allowed:
            raise ValidationError(
                {
                    "non_field_errors": [
                        "Export is not enabled for this resource. Set "
                        "`export_fields` on the viewset to enable it."
                    ]
                }
            )

        requested = request.query_params.get("fields")
        if requested:
            fields = [f.strip() for f in requested.split(",") if f.strip()]
            forbidden = sorted(set(fields) - set(allowed))
            if forbidden:
                raise ValidationError(
                    {"fields": [f"Not exportable: {', '.join(forbidden)}."]}
                )
        else:
            fields = allowed

        export_format = request.query_params.get("file_format", "csv").lower()
        headers = [name.replace("_", " ").replace(".", " ").title() for name in fields]

        queryset = self.filter_queryset(self.get_queryset())
        # Fetch one extra row purely to detect truncation.
        rows = list(queryset[: self.max_export_rows + 1])
        truncated = len(rows) > self.max_export_rows
        rows = rows[: self.max_export_rows]

        model = self.get_model_class()
        basename = model._meta.model_name

        if export_format == "csv":
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(headers)
            for obj in rows:
                writer.writerow(
                    [self._resolve_export_value(obj, name) for name in fields]
                )
            response = HttpResponse(buffer.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{basename}.csv"'

        elif export_format == "xlsx":
            if not XLSX_AVAILABLE:
                raise ValidationError(
                    {
                        "file_format": [
                            "XLSX export requires openpyxl. Install it or use "
                            "file_format=csv."
                        ]
                    }
                )
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = basename[:31]  # Excel caps sheet names at 31 chars.
            sheet.append(headers)
            for obj in rows:
                sheet.append(
                    [self._resolve_export_value(obj, name) for name in fields]
                )
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
            )
            response["Content-Disposition"] = f'attachment; filename="{basename}.xlsx"'

        else:
            raise ValidationError({"file_format": ["Use 'csv' or 'xlsx'."]})

        # Silent truncation reads as a complete export. Say so in a header the
        # client can surface.
        response["X-Export-Row-Count"] = str(len(rows))
        response["X-Export-Truncated"] = "true" if truncated else "false"
        if truncated:
            response["X-Export-Row-Limit"] = str(self.max_export_rows)
        return response

    def _read_upload(self, uploaded_file) -> list[dict]:
        """Parse a CSV or XLSX upload into a list of ``{column: value}``."""
        name = uploaded_file.name.lower()

        def normalize(header):
            if header is None:
                return None
            return str(header).strip().lower().replace(" ", "_")

        if name.endswith(".csv"):
            try:
                text = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError:
                raise ValidationError(
                    {"file": ["The file is not valid UTF-8 encoded text."]}
                )
            reader = csv.reader(io.StringIO(text))
            try:
                header_row = next(reader)
            except StopIteration:
                raise ValidationError({"file": ["The file is empty."]})
            columns = [normalize(h) for h in header_row]
            return [
                dict(zip(columns, row))
                for row in reader
                if any(str(cell).strip() for cell in row)
            ]

        if name.endswith(".xlsx"):
            if not XLSX_AVAILABLE:
                raise ValidationError(
                    {"file": ["XLSX import requires openpyxl. Upload a CSV instead."]}
                )
            try:
                workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            except Exception as exc:  # noqa: BLE001 - openpyxl raises broadly
                raise ValidationError({"file": [f"Unreadable spreadsheet: {exc}"]})

            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                raise ValidationError({"file": ["The file is empty."]})
            columns = [normalize(h) for h in header_row]
            return [
                dict(zip(columns, row)) for row in rows_iter if any(v is not None for v in row)
            ]

        raise ValidationError({"file": ["Upload a .csv or .xlsx file."]})

    @extend_schema(
        summary="Import from CSV or XLSX",
        description=(
            "Rows are validated by the resource serializer. Nothing is "
            "written unless every row validates — a half-imported "
            "spreadsheet is worse than a rejected one."
        ),
        request=FileUploadSerializer,
        responses={200: OpenApiResponse(description="Import result")},
    )
    @action(detail=False, methods=["post"], url_path="bulk_import")
    def bulk_import(self, request):
        self._require_bulk_enabled()

        upload = FileUploadSerializer(data=request.data)
        upload.is_valid(raise_exception=True)
        rows = self._read_upload(upload.validated_data["file"])

        if not rows:
            raise ValidationError({"file": ["No data rows found."]})
        if len(rows) > self.max_bulk_rows:
            raise ValidationError(
                {"file": [f"At most {self.max_bulk_rows} rows per import."]}
            )

        allowed = set(self.import_fields) if self.import_fields else None

        # Validate everything first, then write. The source implementation
        # opened one transaction and validated row-by-row inside it, reporting
        # per-row errors while committing the good rows — so a spreadsheet
        # with one bad row left the database in a state nobody asked for and
        # no obvious way to resume. All-or-nothing is easier to reason about
        # and easier to retry.
        prepared = []
        errors = []

        for index, raw_row in enumerate(rows, start=2):  # row 1 is the header
            row = {k: v for k, v in raw_row.items() if k}
            if allowed is not None:
                row = {k: v for k, v in row.items() if k in allowed}

            serializer = self.get_serializer(data=row)
            if serializer.is_valid():
                prepared.append(serializer)
            else:
                errors.append({"row": index, "errors": serializer.errors})

        if errors:
            return Response(
                {
                    "imported": 0,
                    "total_rows": len(rows),
                    "errors": errors,
                    "message": (
                        f"{len(errors)} row(s) failed validation. Nothing was "
                        f"imported."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        audit = self._audit_kwargs_for_create()
        with transaction.atomic():
            for serializer in prepared:
                serializer.save(**audit)

        return Response(
            {
                "imported": len(prepared),
                "total_rows": len(rows),
                "errors": [],
                "message": f"Imported {len(prepared)} row(s).",
            }
        )

    @extend_schema(
        summary="Download an import template",
        parameters=[
            OpenApiParameter(
                name="file_format",
                type=OpenApiTypes.STR,
                description="'csv' or 'xlsx' (default 'csv').",
            )
        ],
        responses={200: OpenApiResponse(description="Template file")},
    )
    @action(detail=False, methods=["get"], url_path="bulk_import_template")
    def bulk_import_template(self, request):
        """
        Emit a header row derived from the *create serializer*.

        Deriving it from the serializer rather than the model matters: the
        serializer is what will validate the upload, so a template built from
        model fields can list columns the API rejects and omit ones it
        requires (write-only fields, fields the serializer renames).
        """
        model = self.get_model_class()
        serializer = self.get_serializer()

        required, optional, help_text = [], [], {}
        allowed = set(self.import_fields) if self.import_fields else None

        for name, field in serializer.fields.items():
            if field.read_only:
                continue
            if allowed is not None and name not in allowed:
                continue

            text = getattr(field, "help_text", "") or ""
            if not text:
                try:
                    text = getattr(model._meta.get_field(name), "help_text", "") or ""
                except Exception:  # noqa: BLE001 - serializer-only field
                    text = ""
            help_text[name] = str(text)

            (required if field.required else optional).append(name)

        fields = required + optional
        if not fields:
            raise ValidationError(
                {"non_field_errors": ["No importable fields on this resource."]}
            )

        headers = [name.replace("_", " ").title() for name in fields]
        notes = [
            f"{name.replace('_', ' ').title()} "
            f"[{'REQUIRED' if name in required else 'OPTIONAL'}]"
            f"{': ' + help_text[name] if help_text[name] else ''}"
            for name in fields
        ]

        export_format = request.query_params.get("file_format", "csv").lower()
        basename = f"{model._meta.model_name}_import_template"

        if export_format == "csv":
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(headers)
            writer.writerow([])
            writer.writerow([])
            writer.writerow(["# Column reference — delete these rows before importing"])
            for note in notes:
                writer.writerow([f"# {note}"])
            response = HttpResponse(buffer.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{basename}.csv"'
            return response

        if export_format == "xlsx":
            if not XLSX_AVAILABLE:
                raise ValidationError(
                    {"file_format": ["XLSX requires openpyxl. Use file_format=csv."]}
                )
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Import"
            sheet.append(headers)

            # Notes go on a second sheet so the data sheet stays a clean
            # header + rows grid — a user can fill it in and upload it back
            # without deleting anything.
            notes_sheet = workbook.create_sheet("Reference")
            notes_sheet.append(["Column", "Requirement", "Notes"])
            for name in fields:
                notes_sheet.append(
                    [
                        name.replace("_", " ").title(),
                        "REQUIRED" if name in required else "OPTIONAL",
                        help_text[name],
                    ]
                )

            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
            )
            response["Content-Disposition"] = f'attachment; filename="{basename}.xlsx"'
            return response

        raise ValidationError({"file_format": ["Use 'csv' or 'xlsx'."]})

    # ------------------------------------------------------------------
    # Statistics
    # ------------------------------------------------------------------

    @extend_schema(
        summary="Counts for the current filter",
        responses={200: OpenApiResponse(description="Statistics")},
    )
    @action(detail=False, methods=["get"])
    def statistics(self, request):
        queryset = self.get_queryset()
        filtered = self.filter_queryset(queryset)

        stats = {
            "total_count": queryset.count(),
            "filtered_count": filtered.count(),
        }
        stats.update(self.get_extra_statistics(request, queryset, filtered))
        return Response(stats)


class ReadOnlyBaseModelViewSet(SoftDeleteQueryMixin, viewsets.ReadOnlyModelViewSet):
    """
    List/retrieve only, with the same filtering and audit annotations.

    Use for reporting endpoints and reference data. Starting from the
    read-only base and adding actions is safer than starting from the full
    ``BaseModelViewSet`` and remembering to remove them.
    """

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    is_schema_request = BaseModelViewSet.is_schema_request
    get_queryset = BaseModelViewSet.get_queryset
